import openpyxl
import re

def extract_from_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb['full data']

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # print(row[1], row[2])
        name = row[1]
        icon_formula = row[2]

        if icon_formula and isinstance(icon_formula, str) and icon_formula.startswith('=_xlfn.IMAGE('):
            url_match = re.search(r'\"(https?://[^\"]+)\"', icon_formula)

            if url_match:
                url = url_match.group(1)
                name = name.replace("'", "''")
                print(f"UPDATE villagers SET photo_url = '{url}' WHERE name = '{name}';")


if __name__ == "__main__":
    extract_from_excel('data/ACNH_Normalized.xlsx')
